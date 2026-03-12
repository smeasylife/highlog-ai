#!/usr/bin/env python3
"""
PDF와 엑셀 파일을 4페이지/시트씩 묶어서 분리하는 스크립트
사용법: python split_files.py
"""

import os
import sys
from pathlib import Path

try:
    import fitz  # PyMuPDF
except ImportError:
    print("PyMuPDF가 설치되지 않았습니다. 'pip install PyMuPDF'로 설치해주세요.")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("openpyxl가 설치되지 않았습니다. 'pip install openpyxl'로 설치해주세요.")
    sys.exit(1)


def split_pdf_by_pages(input_path: Path, output_dir: Path, pages_per_file: int = 4):
    """PDF를 지정된 페이지 수만큼씩 분리하여 여러 파일로 저장"""
    pdf_name = input_path.stem
    pdf_document = fitz.open(input_path)
    total_pages = len(pdf_document)

    file_count = 1
    for start_page in range(0, total_pages, pages_per_file):
        end_page = min(start_page + pages_per_file, total_pages)
        new_pdf = fitz.open()

        # 페이지 복사
        for page_num in range(start_page, end_page):
            new_pdf.insert_pdf(pdf_document, from_page=page_num, to_page=page_num)

        # 저장
        output_path = output_dir / f"{pdf_name}_part{file_count}.pdf"
        new_pdf.save(output_path)
        new_pdf.close()

        print(f"✅ PDF 생성 완료: {output_path.name} (페이지 {start_page + 1}-{end_page})")
        file_count += 1

    pdf_document.close()


def split_excel_by_sheets(input_path: Path, output_dir: Path, sheets_per_file: int = 4):
    """엑셀을 지정된 시트 수만큼씩 분리하여 여러 파일로 저장"""
    excel_name = input_path.stem
    wb = openpyxl.load_workbook(input_path)
    all_sheets = wb.sheetnames
    total_sheets = len(all_sheets)

    file_count = 1
    for start_idx in range(0, total_sheets, sheets_per_file):
        end_idx = min(start_idx + sheets_per_file, total_sheets)
        sheets_to_copy = all_sheets[start_idx:end_idx]

        # 새 워크북 생성
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)  # 기본 시트 삭제

        # 시트 복사
        for sheet_name in sheets_to_copy:
            original_sheet = wb[sheet_name]
            new_sheet = new_wb.create_sheet(title=sheet_name)

            # 셀 데이터 복사
            for row in original_sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    # 스타일 복사 (선택사항)
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.alignment = cell.alignment.copy()
                        new_cell.number_format = cell.number_format

        # 저장
        output_path = output_dir / f"{excel_name}_part{file_count}.xlsx"
        new_wb.save(output_path)
        new_wb.close()

        print(f"✅ 엑셀 생성 완료: {output_path.name} (시트 {start_idx + 1}-{end_idx})")
        file_count += 1

    wb.close()


def main():
    # 현재 작업 디렉토리 기준
    current_dir = Path.cwd()

    # 입력 파일들 위치 (현재 디렉토리의 .pdf 및 .xlsx 파일)
    pdf_files = list(current_dir.glob("*.pdf"))
    excel_files = list(current_dir.glob("*.xlsx")) + list(current_dir.glob("*.xls"))

    # 결과를 저장할 디렉토리 생성
    output_dir = current_dir / "split_files_output"
    output_dir.mkdir(exist_ok=True)
    print(f"📁 출력 디렉토리 생성: {output_dir}")

    if not pdf_files and not excel_files:
        print("⚠️  현재 디렉토리에 PDF나 엑셀 파일이 없습니다.")
        return

    # PDF 처리
    for pdf_file in pdf_files:
        print(f"\n📄 PDF 처리 중: {pdf_file.name}")
        try:
            split_pdf_by_pages(pdf_file, output_dir, pages_per_file=4)
        except Exception as e:
            print(f"❌ PDF 처리 실패: {e}")

    # 엑셀 처리
    for excel_file in excel_files:
        # 이미 split된 파일은 건너뛰기
        if "part" in excel_file.stem:
            continue

        print(f"\n📊 엑셀 처리 중: {excel_file.name}")
        try:
            split_excel_by_sheets(excel_file, output_dir, sheets_per_file=4)
        except Exception as e:
            print(f"❌ 엑셀 처리 실패: {e}")

    print(f"\n✨ 모든 작업 완료! 결과는 {output_dir} 디렉토리에 저장되었습니다.")


if __name__ == "__main__":
    main()
